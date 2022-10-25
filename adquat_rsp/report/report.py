# Copyright 2017 Creu Blanca
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl.html).

from odoo import api, fields, models
from xlrd import open_workbook
from odoo.modules.module import get_module_resource
import base64
from io import BytesIO
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

class ProjectXlsx(models.AbstractModel):
    _name = "report.adquat_rsp.report_document_vt_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Project XLSX Report"

    def generate_xlsx_report(self, workbook, data, projects):
        projects.ensure_one()
        user_date_format = self.env['res.lang']._lang_get(self.env.user.lang).date_format
        ws = workbook.active

        if projects.date_vt:
            ws.cell(3,3).value = projects.date_vt.strftime(user_date_format)
        if projects.tech_id:
            ws.cell(3,6).value = projects.tech_id.name

        if len(projects.name_partner.split(' ')) > 1:
            ws.cell(6,3).value = projects.name_partner.split(' ')[1]
            ws.cell(6,6).value = projects.name_partner.split(' ')[0]
        else:
            ws.cell(6,3).value = projects.name_partner

        ws.cell(7,3).value = (projects.street or '') + (projects.street2 and '\n' + projects.street2 or '')
        ws.cell(7,6).value = projects.birth_partner
        ws.cell(8,3).value = projects.phone_partner
        ws.cell(8,6).value = projects.mail_partner
        ws.cell(9,3).value = projects.partner_id and projects.partner_id.mobile or ''

    def create_xlsx_report(self, docids, data):
        objs = self._get_objs_for_report(docids, data)
        #NEW
        # file_data = self.env.ref('adquat_rsp.attachment_document_vt').datas
        xlsx_file_path = get_module_resource('adquat_rsp', 'report', 'document_vt.xlsx')
        workbook = load_workbook(xlsx_file_path)
        self.generate_xlsx_report(workbook, data, objs)
        file_data = BytesIO(save_virtual_workbook(workbook))
        file_data.seek(0)
        return file_data.read(), "xlsx"