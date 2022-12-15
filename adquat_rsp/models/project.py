from datetime import date
from odoo import api, fields, models
from odoo.tools import date_utils
from openpyxl import load_workbook, Workbook
from odoo.modules.module import get_module_resource
from io import BytesIO
from openpyxl.writer.excel import save_virtual_workbook
import base64
import io
import json
import base64
from odoo.exceptions import UserError

class ProjectProject(models.Model):
    _name = 'project.project'
    _inherit = ['mail.thread.phone', 'project.project']
    def button_dummy(self):
        # TDE FIXME: this button is very interesting
        return True

    def _get_document_partner(self):
        return self.partner_id
    def _phone_get_number_fields(self):
        """ This method returns the fields to use to find the number to use to
        send an SMS on a record. """
        return ['phone_partner']

    #HERITAGE
    partner_id = fields.Many2one('res.partner', string='Client', auto_join=True, tracking=True, required=True,
                                 domain=lambda self: [('category_id', 'in', self.env.ref('adquat_rsp.res_partner_category_customer').id),
                                                      '|', ('company_id', '=', False), ('company_id', '=', self.env.company)])

## Infos client: Onglet fiche client
    # @api.depends('partner_id', 'partner_id.name')
    # def _compute_partner_name(self):
    #     for project in self:
    #         # partner_name = project.partner_id and project.partner_id.name or ''
            # if partner_name:
            #     split = partner_name.split(' ')
            #     if len(split) > 1:
            #         project.name_partner = split[-1]
            #         project.prenom_partner = split[0]
            #     else:
            #         project.name_partner = partner_name
            #         project.prenom_partner = ''

    # name_partner = fields.Char(string="Nom", compute='_compute_partner_name', store=True)
    name_partner = fields.Char(string="Nom", related='partner_id.lastname')
    # prenom_partner = fields.Char(string="Prénom", compute='_compute_partner_name', store=True)
    prenom_partner = fields.Char(string="Prénom", related='partner_id.firstname')
    birth_partner = fields.Date(string="Date de Naissance", related='partner_id.date_birth_partner')
    street = fields.Char(related='partner_id.street')
    street2 = fields.Char(related='partner_id.street2')
    zip = fields.Char(change_default=True, related='partner_id.zip')
    city = fields.Char(related='partner_id.city')
    state_id = fields.Many2one("res.country.state", string='State', ondelete='restrict',
                               domain="[('country_id', '=?', country_id)]", related='partner_id.state_id')
    country_id = fields.Many2one('res.country', string='Country', ondelete='restrict', related='partner_id.country_id')
    country_code = fields.Char(related='country_id.code', string="Country Code")
    phone_partner = fields.Char(string="Téléphone", related="partner_id.phone")
    mobile_partner = fields.Char(string="Mobile", related="partner_id.mobile")
    mail_partner = fields.Char(string="Email", related="partner_id.email")
    time = fields.Char("Temps de Route")
    parrainage = fields.Many2one('res.partner', string="Parrainage",
                                 domain=lambda self: [('category_id', 'in', self.env.ref('adquat_rsp.res_partner_category_customer').id),
                                '|', ('company_id', '=', False), ('company_id', '=', self.env.company)])

    @api.onchange('partner_id')
    def _onchange_address(self):
        if self.partner_id:
            if not self.street:
                self.street = self.partner_id.street
            if not self.street2:
                self.street2 = self.partner_id.street2
            if not self.city:
                self.city = self.partner_id.city
            if not self.zip:
                self.zip = self.partner_id.zip
            if not self.state_id:
                self.state_id = self.partner_id.state_id.id

## Infos dossier: Onglet fiche client
    existing_power = fields.Float("Puissance existance")
    rv_or_auto = fields.Selection([
        ('rv', 'RV'),
        ('auto', 'AUTO')
    ], string="RV ou AUTO")
    crae = fields.Char("N°CRAE")
    bta = fields.Char('N°BTA')
    msb = fields.Char('Numéro de série MSB')

## Fichier à joindre + infos complémentaires: Ongmet fiche client
    devis_and_chq = fields.Many2many('ir.attachment', 'ir_attachment_devis_chq', string='Devis + chèque')
    cgv = fields.Many2many('ir.attachment', 'ir_attachment_cdv', string='CGV Paraphées')
    taxes_foncieres = fields.Many2many('ir.attachment', 'ir_attachment_taxes_foncieres', string='Taxes Foncières')
    fact_elec = fields.Many2many('ir.attachment', 'ir_attachment_fact_elec', string='Facture électricité')
    mandat_mairie = fields.Many2many('ir.attachment', 'ir_attachment_mandat_mairie', string='Mandat Mairie')
    mandat_enedis = fields.Many2many('ir.attachment', 'ir_attachment_mandat_enedis', string='Mandat Enedis')
    mandat_OA = fields.Many2many('ir.attachment', 'ir_attachment_mandat_oa', string='Mandat OA')
    date_recepisse = fields.Date('Date récépissé mairie')
    abf = fields.Boolean('ABF', default=False)
    domofinance = fields.Boolean('Domofinance', default=False)
    dossier_complet = fields.Boolean('Dossier Complet', default=False, compute="_compute_dossier_complet", readonly=True,
                                     help="Pour qu'un dossier soit complet il faut : \n" \
                                          "* Coordonnées : Nom + Prénom + Rue + Code Postal + Ville + Pays + Téléphone ou Mobile + Email \n" \
                                          "* Informations importantes : Montant HT + Date signature + Puissance choisie + Commerciaux \n" \
                                          "* Si puissance existante : RV/Auto + CRAE + BTA \n" \
                                          "* Les pièces jointes : Si OA, il faut la PJ OA"
                                     )
    vt_complet = fields.Boolean('VT Complete', default=False, compute="_compute_vt_complet", readonly=True)

    @api.depends('name_partner', 'prenom_partner', 'birth_partner', 'street', 'city', 'zip', 'country_id', 'phone_partner',
                 'mobile_partner', 'mail_partner', 'devis_and_chq', 'cgv', 'taxes_foncieres', 'fact_elec', 'mandat_mairie',
                 'mandat_OA', 'gestion_surplus', 'mandat_enedis', 'amount_ht', 'date_signature', 'power_choose', 'user_ids')
    def _compute_dossier_complet(self):
        for project in self:
            project.dossier_complet = False
            #COORDONNEES
            coordonnees_complete = project.name_partner and project.prenom_partner \
                                   and project.street and project.city and project.zip and project.country_id \
                                   and (project.phone_partner or project.mobile_partner) and project.mail_partner
            #POWER
            if project.existing_power:
                power_complete = project.rv_or_auto and project.crae and project.bta or False
            else:
                power_complete = True

            #PJs
            pjs_standard = project.devis_and_chq and project.taxes_foncieres and project.cgv and project.fact_elec \
                           and project.mandat_mairie and project.mandat_enedis and project.amount_ht \
                           and project.date_signature and project.power_choose and project.user_ids

            #CHECK
            if coordonnees_complete and pjs_standard and power_complete:
                if (project.gestion_surplus == 'oa' and project.mandat_OA) or project.gestion_surplus != 'oa':
                    project.dossier_complet = True
            if project.dossier_complet and project.stage_id == self.env.ref('adquat_rsp.project_project_stage_new'):
                project.stage_id = self.env.ref('adquat_rsp.project_project_stage_vt_toplan')

    @api.depends('date_vt', 'tech_id', 'file_to_join', 'pic_to_join')
    def _compute_vt_complet(self):
        for project in self:
            project.vt_complet = False
            if project.date_vt and project.tech_id and project.file_to_join and project.pic_to_join:
                project.vt_complet = True

    @api.depends('amount_ht', 'prct_commission')
    def _compute_commission(self):
        for project in self:
            if project.prct_commission:
                project.amount_commission = project.amount_ht * project.prct_commission
            else:
                project.amount_commission = 0.0
    @api.depends('pose_ids', 'pose_ids.date_start_install','pose_ids.date_end_install', 'pose_id_mylight', 'pose_id_enphase', 'stage_id', 'gestion_surplus')
    def _compute_pose(self):
        for project in self:
            project.pose_id = False
            project.pose_id_good = False
            if project.pose_ids:
                project.pose_id = project.pose_ids[-1].id
                if project.stage_id == self.env.ref('adquat_rsp.project_project_stage_pose_planned'):
                    if project.gestion_surplus == 'msb':
                        project.pose_id_good = project.pose_id_mylight and project.pose_id_enphase or False
                    else:
                        project.pose_id_good = project.pose_id_enphase or False

## Champ hors onglet
    gestion_surplus = fields.Selection([
        ('oa', 'OA'),
        ('msb', 'MSB'),
        ('other', 'Autres')
    ], string="Gestion Surplus", default='oa', required=True)
    currency_id = fields.Many2one('res.currency', related='company_id.currency_id')
    amount_ht = fields.Monetary("Montant HT", group_operator="sum")
    prct_commission = fields.Float("Commission",default=0.1)
    amount_commission = fields.Monetary("Montant Commission", compute="_compute_commission", group_operator="sum", store=True)
    date_signature = fields.Date("Date Signature Commande")
    power_choose = fields.Float("Puissance Choisie")
    date_vt = fields.Datetime("Date et heure VT")
    date_mairie = fields.Date("Date accord mairie")
    pose_id = fields.Many2one('project.pose',string="Pose actuelle", compute="_compute_pose", store=True)
    pose_id_mylight = fields.Boolean(string="MyLight Pose actuelle", related="pose_id.monitoring_mylight")
    pose_id_enphase = fields.Boolean(string="Enphase Pose actuelle", related="pose_id.enphase")
    pose_id_good = fields.Boolean(string="Pose OK", compute="_compute_pose")
    date_start_install = fields.Date("Date de début de l'installation", related="pose_id.date_start_install", store=True)
    date_end_install = fields.Date("Date de fin de l'installation", related="pose_id.date_end_install", store=True)
    #techs_name = fields.Char("Nom des techs")
    tech_ids = fields.Many2many('hr.employee', 'project_tech_ids', string="Techniciens",
                                domain=lambda self: [('department_id', '=', self.env.ref('adquat_rsp.hr_department_tech').id)])
    user_ids = fields.Many2many('hr.employee', 'project_user_ids', string="Commerciaux",
                                domain=lambda self: [('department_id', '=', self.env.ref('hr.dep_sales').id)])
    date_mise_service_enedis = fields.Date('Date de mise en service Enedis')


## fichier et infos onglet VT
    tech_id = fields.Many2one('hr.employee', string='Technicien VT',
                              domain=lambda self: [('department_id', '=', self.env.ref('adquat_rsp.hr_department_tech').id)])
    file_to_join = fields.Many2many('ir.attachment', 'ir_attachment_file_join', string='Fichiers à joindre')
    pic_to_join = fields.Many2many('ir.attachment', 'ir_attachment_pic_join', string='Photos à joindre')
    vt_file = fields.Many2one('ir.attachment', string='Fiche Technique générée', copy=False)
    # vt_filed = fields.Many2many('ir.attachment', 'ir_attachment_project_vt', string='Fiche technique remplie')
    # vt_filename = fields.Char("VT Filename")

    @api.onchange('tech_id', 'file_to_join', 'pic_to_join', 'date_vt')
    def _on_change_stage_id_vt(self):
        for project in self:
            if project.date_vt:
                if project.file_to_join and project.pic_to_join and project.tech_id and project.stage_id.id == self.env.ref('adquat_rsp.project_project_stage_vt_planned').id:
                    project.stage_id = self.env.ref('adquat_rsp.project_project_stage_mairie_todo').id
                else:
                    project.stage_id = self.env.ref('adquat_rsp.project_project_stage_vt_planned').id
            else:
                pass

## fichiers et infos onglet Mairie
    done = fields.Boolean('Faite / Pas Faite')
    sending_date_mairie = fields.Date('Date d\'envoi ')
    mairie_answer_sent = fields.Boolean('Réponse de la mairie envoyée au client')
    mairie_answer_date = fields.Date('Date de réponse')
    mairie_answer = fields.Selection([
        ('yes', 'Accord'),
        ('no', 'Refus')
    ], string="Réponse")
    mairie_answer_to_join = fields.Many2many('ir.attachment', 'ir_attachment_mairie_answer', string='Accord/Refus à importer')
    recepisse_to_join = fields.Many2many('ir.attachment', 'ir_attachment_recepisse', string='Récépissé fichier')
    other_attachments_to_join = fields.Many2many('ir.attachment', 'ir_attachment_other', string='Pièces complémentaires')
    abf_to_join = fields.Many2many('ir.attachment', 'ir_attachment_abf', string='ABF')
    rsp_to_join = fields.Many2many('ir.attachment', 'ir_attachment_rsp', string='Décharge RSP')

    @api.onchange('sending_date_mairie')
    def _onchange_done(self):
        for project in self:
            if project.sending_date_mairie:
                project.done = True
                project.stage_id = self.env.ref('adquat_rsp.project_project_stage_mairie_done').id
            else:
                project.done = False

    @api.onchange('mairie_answer', 'mairie_answer_to_join', 'rsp_to_join', 'mairie_answer_sent')
    def _onchange_stage_id_mairie(self):
        for project in self:
            if project.mairie_answer:
                project.mairie_answer_date = fields.Date.today()
                project.date_mairie = fields.Date.today()
            if ((project.mairie_answer == 'yes' and project.mairie_answer_to_join) or project.rsp_to_join) and project.stage_id.id == self.env.ref('adquat_rsp.project_project_stage_mairie_done').id:
                project.stage_id = self.env.ref('adquat_rsp.project_project_stage_pose_toplan').id
            if project.mairie_answer == 'yes' and not project.mairie_answer_sent and project.mairie_answer_to_join:
                args = {
                    'auto_delete_message': True,
                    'subtype_id': self.env['ir.model.data']._xmlid_to_res_id('mail.mt_note'),
                    'email_layout_xmlid': 'mail.mail_notification_light'
                }
                if self.env.ref('adquat_rsp.mail_auto_accord_mairie'):
                    project.message_post_with_template(self.env.ref('adquat_rsp.mail_auto_accord_mairie').id,  **{
                        'auto_delete_message': False,
                        'subtype_id': self.env['ir.model.data']._xmlid_to_res_id('mail.mt_note'),
                        'email_layout_xmlid': 'mail.mail_notification_light',
                        'attachment_ids': project.mairie_answer_to_join and project.mairie_answer_to_join.ids or False,
                    })
                    project.mairie_answer_sent = True
            else:
                pass
    def action_fsm_navigate(self):
        if not self.partner_id.partner_latitude and not self.partner_id.partner_longitude:
            self.partner_id.geo_localize()
        url = "https://www.google.com/maps/dir/?api=1&origin=%s,%s&destination=%s,%s" % (self.company_id.partner_id.partner_latitude, self.company_id.partner_id.partner_longitude, self.partner_id.partner_latitude, self.partner_id.partner_longitude)
        return {
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'new'
        }

##Fichiers et infos onglet Pose
    pose_ids = fields.One2many('project.pose', 'project_id')
    # return_caution = fields.Boolean('Retour chq Caution', default=False)
    #attachment_pose_ids = fields.Many2many('ir.attachment', string='Pose Attachments', compute='_compute_attachment_pose_ids', store=True)
    # aft = fields.Many2many('ir.attachment', 'ir_attachment_all_aft', string='AFT')
    # picture = fields.Many2many('ir.attachment', 'ir_attachment_all_picture', string='Photos')
    # calepinage_emphase = fields.Many2many('ir.attachment', 'ir_attachment_all_calepinage', string='Calepinage Enphase')
    # implantation_emphase = fields.Many2many('ir.attachment', 'ir_attachment_all_implantation', string='Rapport Enphase')
    # quotation_alaska = fields.Many2many('ir.attachment', 'ir_attachment_all_quot_alaska', string='Devis Alaska')
    # invoice_alaska = fields.Many2many('ir.attachment', 'ir_attachment_all_inv_alaska', string='Facture alaska')
    # invoice_finalRsp = fields.Many2many('ir.attachment', 'ir_attachment_all_inv_rsp', string='Facture finale RSP client')
    finish_pose_display = fields.Boolean('Pose terminée affichée', compute='_compute_finish_pose_display')
    has_complete_partner_address = fields.Boolean(compute='_compute_has_complete_partner_address')
    @api.depends('name_partner', 'prenom_partner', 'birth_partner', 'street', 'city', 'zip', 'country_id', 'phone_partner',
                 'mobile_partner', 'mail_partner')
    def _compute_has_complete_partner_address(self):
        for project in self:
            project.has_complete_partner_address = project.name_partner and project.prenom_partner \
                                   and project.street and project.city and project.zip and project.country_id \
                                   and (project.phone_partner or project.mobile_partner) and project.mail_partner or False

    @api.onchange('date_start_install','date_end_install')
    def _onchange_stage_id(self):
        for project in self:
            if project.date_start_install and project.date_end_install and project.stage_id.id == self.env.ref('adquat_rsp.project_project_stage_pose_toplan').id:
                project.stage_id = self.env.ref('adquat_rsp.project_project_stage_pose_planned').id
            else:
                pass

    def finish_pose(self):
        self.stage_id = self.env.ref('adquat_rsp.project_project_stage_mes').id

    def create_fdi(self):
        self.ensure_one()
        if self.fdi_ids.filtered(lambda fdi: fdi.state == 'planif'):
            raise UserError("Il y a déjà une FDI programmée. \nVeuillez la clore avant d'en créer une autre.")

        new_context = self.env.context.copy()
        new_context['default_type'] = 'fdi'
        new_context['default_project_id'] = self.id

        return {
            'name': 'Assistant FDI',
            'view_mode': 'form',
            'res_model': 'project.fdi.sav.wizard',
            'type': 'ir.actions.act_window',
            'target': 'new',
            'context': new_context,
        }

## Infos FDI
    fdi_id = fields.Many2one('fdi.object',string="Dernière FDI", compute="_compute_fdi", store=True)
    fdi_ids = fields.One2many('fdi.object', 'project_id')
    aft_file = fields.Many2one('ir.attachment', string='AFT générée', copy=False)
    date_fdi = fields.Datetime("Date FDI", related="fdi_id.date", store=True, tracking=True)
    state_fdi = fields.Selection([('planif', 'Programmée'),('finish', 'Terminée'),],
                                 string="Etat FDI", related="fdi_id.state", store=True)

    def _track_template(self, changes):
        res = super(ProjectProject, self)._track_template(changes)
        project = self[0]
        if any(field in ('date_fdi', 'date_sav') for field in changes) and 'stage_id' not in changes and project.stage_id.mail_template_id:
            res['stage_id'] = (project.stage_id.mail_template_id, {
                'auto_delete_message': False,
                'subtype_id': self.env['ir.model.data']._xmlid_to_res_id('mail.mt_note'),
                'email_layout_xmlid': 'mail.mail_notification_light'
            }
        )
        return res
    @api.depends('fdi_ids', 'fdi_ids.date')
    def _compute_fdi(self):
        for project in self:
            last_fdi = project.fdi_ids[-1] if project.fdi_ids else False
            if last_fdi and last_fdi.date:
                project.fdi_id = last_fdi.id
            else:
                project.fdi_id = False

## Infos SAV
    sav_ids = fields.One2many('sav.object', 'project_id')
    sav_id = fields.Many2one('sav.object',string="Dernièr SAV", compute="_compute_sav", store=True)
    sav_file = fields.Many2one('ir.attachment', string='SAV généré', copy=False)
    date_sav = fields.Datetime("Date SAV", related="sav_id.date", store=True, tracking=True)
    state_sav = fields.Selection([
        ('planif', 'Programmé'),
        ('finish', 'Terminé'),
    ], string="Etat SAV", related="sav_id.state", store=True)

    @api.depends('sav_ids', 'sav_ids.date')
    def _compute_sav(self):
        for project in self:
            last_sav = project.sav_ids[-1] if project.sav_ids else False
            if last_sav and last_sav.date:
                project.sav_id = last_sav.id
            else:
                project.sav_id = False

## Infos Enedis et Consuel: Onglet mise en servcie
    # Enedis
    numb_pdr = fields.Char('Créat° Numéro PDR')
    consuel_transmitted_enedis = fields.Boolean('Consuel transmis à Enedis', default=False)
    synthese = fields.Many2many('ir.attachment', 'ir_attachment_synthese_enedis', string='Synthèse')
    enedis_done = fields.Boolean('Enedis validé')

    @api.onchange('numb_pdr', 'synthese')
    def _onchange_enedis_done(self):
        if self.numb_pdr and self.synthese:
            self.enedis_done = True
            template = self.env.ref('adquat_rsp.mail_auto_synthese_enedis')
            if template:
                self.message_post_with_template(template.id,  **{
                    'auto_delete_message': False,
                    'subtype_id': self.env['ir.model.data']._xmlid_to_res_id('mail.mt_note'),
                    'email_layout_xmlid': 'mail.mail_notification_light',
                    'attachment_ids': [(6, 0, self.synthese.ids)]})
        else:
            self.enedis_done = False

    def finish_project(self):
        self.stage_id = self.env.ref('adquat_rsp.project_project_stage_done').id

    def create_sav(self):
        self.ensure_one()
        if self.sav_ids.filtered(lambda sav: sav.state == 'planif'):
            raise UserError("Il y a déjà un SAV programmé. \nVeuillez le clore avant d'en créer un autre.")
        new_context = self.env.context.copy()
        new_context['default_type'] = 'sav'
        new_context['default_project_id'] = self.id
        return {
            'name': 'Assistant SAV',
            'view_mode': 'form',
            'res_model': 'project.fdi.sav.wizard',
            'type': 'ir.actions.act_window',
            'target': 'new',
            'context': new_context,
        }

    # Consuel
    shipping_number = fields.Char('Numéro d\'envoi')
    type_of_visit = fields.Selection([
        ('1', 'AUDIT'),
        ('2', 'CONSUEL')
    ], string="Type de Visite")
    intended_date = fields.Date('Date prévue')
    date_contre_visite = fields.Date('Date contre-visite')
    date_attestation = fields.Date('Date attestation visée')
    pdf_consuel = fields.Many2many('ir.attachment', 'ir_attachment_pdf_consuel', string='PDF du Consuel')
    fileTech_and_schema = fields.Many2many('ir.attachment', 'ir_attachment_filetech', string='Doss Tech + Schéma')
    consuel_done = fields.Boolean('Consuel validé')
    consuel_sent = fields.Boolean('Consuel envoyé')

    #MSB
    contrat_mylight = fields.Boolean('Contrat MyLight', default=False)

    @api.onchange('contrat_mylight')
    def _onchange_contrat_mylight(self):
        if self.contrat_mylight:
            template = self.env.ref('adquat_rsp.mail_auto_end_install_souscription_mylight')
            self.message_post_with_template(template.id,  **{
                'auto_delete_message': False,
                'subtype_id': self.env['ir.model.data']._xmlid_to_res_id('mail.mt_note'),
                'email_layout_xmlid': 'mail.mail_notification_light',})

    @api.onchange('shipping_number', 'fileTech_and_schema')
    def _onchange_consuel_done(self):
        if self.shipping_number and self.fileTech_and_schema:
            self.consuel_done = True
        else:
            self.consuel_done = False
    @api.onchange('pdf_consuel')
    def _onchange_pdf_consuel(self):
        if self.pdf_consuel and not self.consuel_sent:
            template = self.gestion_surplus == 'msb' and self.env.ref('adquat_rsp.mail_auto_envoi_consuel_if_msb') or \
                self.env.ref('adquat_rsp.mail_auto_envoi_consuel_if_oa')

            self.message_post_with_template(template.id,  **{
                'auto_delete_message': False,
                'subtype_id': self.env['ir.model.data']._xmlid_to_res_id('mail.mt_note'),
                'email_layout_xmlid': 'mail.mail_notification_light',
                'attachment_ids': [(6, 0, self.pdf_consuel.ids + [self.env.ref('adquat_rsp.attachment_oa_bien_signer').id,
                                                      self.env.ref('adquat_rsp.attachment_oa_modifier_numero').id,
                                                      self.env.ref('adquat_rsp.attachment_oa_recuperer_mdp').id,
                                                      self.env.ref('adquat_rsp.attachment_oa_livret_producteur').id])]})

    nb_quotation_validate = fields.Integer('Devis validés', default=0)
    nb_vt_to_planif = fields.Integer('VT à planifier', default=0)
    nb_pose_to_planif = fields.Integer('Pose à planifier', default=0)
    nb_pose_planif = fields.Integer('Pose Finies', default=0)
    nb_sav_finish = fields.Integer('SAV finis', default=0)
    nb_sav_to_planif = fields.Integer('SAV à planifier', default=0)
    nb_fdi_to_planif = fields.Integer('FDI à planifier', default=0)
    nb_fdi_finish = fields.Integer('FDI finies', default=0)
    nb_project_finish = fields.Integer('Dossiers clôturés', default=0)

    def action_generate_xls(self):
        self.ensure_one()
        xlsx_file_path = get_module_resource('adquat_rsp', 'report', 'document_vt.xlsx')
        workbook = load_workbook(xlsx_file_path)
        user_date_format = self.env['res.lang']._lang_get(self.env.user.lang).date_format
        ws = workbook.active

        if self.date_vt:
            ws.cell(3, 3).value = self.date_vt.strftime(user_date_format)
        if self.tech_id:
            ws.cell(3, 6).value = self.tech_id.name
        if self.name_partner:
            ws.cell(6, 3).value = self.name_partner
        if self.prenom_partner:
            ws.cell(6, 6).value = self.prenom_partner

        ws.cell(7, 3).value = (self.street or '') + (self.street2 and '\n' + self.street2 or '') + \
                              ('\n' + self.zip or '') + (' ' + self.city or '')
        ws.cell(7, 6).value = self.birth_partner
        ws.cell(8, 3).value = self.phone_partner
        ws.cell(8, 6).value = self.mail_partner
        ws.cell(9, 3).value = self.partner_id and self.partner_id.mobile or ''
        file_data = BytesIO(save_virtual_workbook(workbook))
        file_data.seek(0)
        file_data = base64.encodebytes(file_data.read())
        if file_data:
            if self.vt_file:
                self.vt_file.write({'datas':file_data})
                vt = self.vt_file
            else:
                vt = self.env['ir.attachment'].create({
                    'name': 'Visite Technique %s %s.xlsx' % (self.partner_id.name, self.name),
                    'datas': file_data,
                    'type': 'binary',
                    'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'res_id': self.id,
                    'res_model':'project.project',
                })
                self.vt_file = vt
            if vt:
                base_url = self.env['ir.config_parameter'].get_param('web.base.url')
                download_url = '/web/content/' + str(vt.id) + '?download=true'
                return {
                    "type": "ir.actions.act_url",
                    "url": str(base_url) + str(download_url),
                    "target": "new",
                }
    def action_generate_xls_aft(self):
        self.ensure_one()
        xlsx_file_path = get_module_resource('adquat_rsp', 'report', 'aft.xlsx')
        workbook = load_workbook(xlsx_file_path)
        user_date_format = self.env['res.lang']._lang_get(self.env.user.lang).date_format
        ws = workbook.active

        if self.name_partner:
            ws.cell(5, 2).value = self.name_partner
        if self.prenom_partner:
            ws.cell(5, 6).value = self.prenom_partner

        ws.cell(6, 2).value = (self.street or '') + (self.street2 and '\n' + self.street2 or '')
        ws.cell(7, 2).value = self.zip or ''
        ws.cell(7, 6).value = self.city or ''
        ws.cell(8, 2).value = self.mail_partner
        ws.cell(8, 6).value = self.mobile_partner or self.phone_partner or ''

        file_data = BytesIO(save_virtual_workbook(workbook))
        file_data.seek(0)
        file_data = base64.encodebytes(file_data.read())
        if file_data:
            if self.aft_file:
                self.aft_file.write({'datas': file_data})
                vt = self.aft_file
            else:
                vt = self.env['ir.attachment'].create({
                    'name': 'AFT %s %s.xlsx' % (self.partner_id.name, self.name),
                    'datas': file_data,
                    'type': 'binary',
                    'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'res_id': self.id,
                    'res_model': 'project.project',
                })
                self.aft_file = vt
            if vt:
                base_url = self.env['ir.config_parameter'].get_param('web.base.url')
                download_url = '/web/content/' + str(vt.id) + '?download=true'
                return {
                    "type": "ir.actions.act_url",
                    "url": str(base_url) + str(download_url),
                    "target": "new",
                }
    def action_generate_xls_sav(self):
        self.ensure_one()
        xlsx_file_path = get_module_resource('adquat_rsp', 'report', 'sav.xlsx')
        workbook = load_workbook(xlsx_file_path)
        user_date_format = self.env['res.lang']._lang_get(self.env.user.lang).date_format
        ws = workbook.active
        #
        if self.date_sav:
            ws.cell(3, 2).value = self.date_sav.strftime(user_date_format)
        if self.tech_id:
            ws.cell(3, 6).value = self.tech_id.name
        if self.name_partner:
            ws.cell(6, 3).value = self.name_partner
        if self.prenom_partner:
            ws.cell(6, 4).value = self.prenom_partner

        ws.cell(7, 2).value = (self.street or '') + (self.street2 and '\n' + self.street2 or '')
        ws.cell(8, 2).value = self.zip or ''
        ws.cell(8, 6).value = self.city or ''
        ws.cell(9, 2).value = self.mail_partner or ''
        ws.cell(10, 2).value = self.mobile_partner or ''
        ws.cell(10, 6).value = self.phone_partner or ''

        file_data = BytesIO(save_virtual_workbook(workbook))
        file_data.seek(0)
        file_data = base64.encodebytes(file_data.read())
        if file_data:
            if self.sav_file:
                self.sav_file.write({'datas': file_data})
                vt = self.sav_file
            else:
                vt = self.env['ir.attachment'].create({
                    'name': 'SAV %s %s.xlsx' % (self.partner_id.name, self.name),
                    'datas': file_data,
                    'type': 'binary',
                    'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'res_id': self.id,
                    'res_model': 'project.project',
                })
                self.sav_file = vt
            if vt:
                base_url = self.env['ir.config_parameter'].get_param('web.base.url')
                download_url = '/web/content/' + str(vt.id) + '?download=true'
                return {
                    "type": "ir.actions.act_url",
                    "url": str(base_url) + str(download_url),
                    "target": "new",
                }
    #DOCUMENTS
    use_subfolders = fields.Boolean("Création d'un sous-dossier par onglet", default=True,
                                   help='Crée un sous-dossier par onglet afin de faciliter le classement')
    documents_folder_fiche = fields.Many2one('documents.folder', string="Fiche Client",
                                          domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]",
                                          copy=False)
    documents_folder_vt = fields.Many2one('documents.folder', string="Visite Technique",
                                             domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]",
                                             copy=False)
    documents_folder_mairie = fields.Many2one('documents.folder', string="Mairie",
                                             domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]",
                                             copy=False)
    documents_folder_pose = fields.Many2one('documents.folder', string="Pose",
                                             domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]",
                                             copy=False)
    documents_folder_fdi = fields.Many2one('documents.folder', string="FDI",
                                             domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]",
                                             copy=False)
    documents_folder_mes = fields.Many2one('documents.folder', string="Mise en Service",
                                             domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]",
                                             copy=False)
    documents_folder_sav = fields.Many2one('documents.folder', string="SAV",
                                             domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]",
                                             copy=False)

    def _get_document_folder(self):
        return self.documents_folder_id
    def _get_subfolders_info(self):
        return {(0, 'Fiche Client'): {'folder_field':'documents_folder_fiche',
                    'fields':['devis_and_chq','cgv','taxes_foncieres','fact_elec','mandat_mairie','mandat_enedis']},
                (1, 'Visite Technique'): {'folder_field':'documents_folder_vt',
                    'fields':['file_to_join','pic_to_join','vt_file']},
                (2, 'Mairie'): {'folder_field':'documents_folder_mairie',
                    'fields':['mairie_answer_to_join','recepisse_to_join','other_attachments_to_join','abf_to_join','rsp_to_join']},
                (3, 'Pose'): {'folder_field':'documents_folder_pose',
                    'fields':['aft','picture','calepinage_emphase','implantation_emphase','quotation_alaska','invoice_alaska','invoice_finalRsp']},
                (4, 'FDI'): {'folder_field':'documents_folder_fdi',
                    'fields':['aft_file']},
                (5, 'Mise en Service'): {'folder_field':'documents_folder_mes',
                    'fields':['synthese','pdf_consuel','fileTech_and_schema']},
                (6, 'SAV'): {'folder_field':'documents_folder_sav',
                    'fields':['aft_file']},
        }
    @api.model_create_multi
    def create(self, vals_list):
        projects = super().create(vals_list)
        if not self.env.context.get('no_create_folder'):
            projects.filtered(lambda project: project.use_documents)._create_missing_subfolders()
        return projects
    def write(self, vals):
        res = super(ProjectProject,self).write(vals)
        fdi_change = 'fdi_ids' in vals and any(fdi[2] and fdi[2].get('date',False) for fdi in vals['fdi_ids']) or False
        sav_change = 'sav_ids' in vals and any(sav[2] and sav[2].get('date',False) for sav in vals['sav_ids']) or False
        if fdi_change or sav_change:
            self._send_sms()
        if vals.get('use_documents'):
            self._create_missing_subfolders()
        if vals.get('name'):
            for project in self.filtered(lambda p: p.documents_folder_id):
                project.documents_folder_id.name = vals['name']

        TAB_DIC = self._get_subfolders_info()
        for project in self:
            for field_info in TAB_DIC.values():
                folder_to_change = list(set(field_info['fields']) & set(vals))
                new_subfolder_field = field_info['folder_field']
                for field_tmp in folder_to_change:
                    attachment = project[field_tmp]
                    #LA ON CHERCHER LE DOCUMENT DE LA PJ POUR MODIFIER
                    document = self.env['documents.document'].search([('attachment_id', 'in',attachment.ids)])
                    new_subfolder = project[new_subfolder_field]
                    if document and new_subfolder:
                        document.folder_id = new_subfolder.id
        return res
    @api.returns('self', lambda value: value.id)
    def copy(self, default=None):
        project = super(ProjectProject).copy(default)
        if not self.env.context.get('no_create_folder') and project.use_subfolders and self.documents_folder_id:
            project.documents_folder_fiche = self.documents_folder_fiche.copy({'name': project.name})
            project.documents_folder_vt = self.documents_folder_vt.copy({'name': project.name})
            project.documents_folder_mairie = self.documents_folder_mairie.copy({'name': project.name})
            project.documents_folder_pose = self.documents_folder_pose.copy({'name': project.name})
            project.documents_folder_sav = self.documents_folder_sav.copy({'name': project.name})
            project.documents_folder_fdi = self.documents_folder_fdi.copy({'name': project.name})
            project.documents_folder_mes = self.documents_folder_mes.copy({'name': project.name})
        return project
    def _create_missing_subfolders(self):
        TAB_DIC = self._get_subfolders_info()
        for project in self:
            created_folders = []
            if project.use_subfolders and project.documents_folder_id:
                for (seq, tab), field_info in TAB_DIC.items():
                    field = field_info.get('folder_field',False)
                    if field and not project[field]:
                        folder_vals = {
                            'sequence':seq,
                            'name': tab,
                            'parent_folder_id': project.documents_folder_id.id,
                            'company_id': project.company_id.id,
                        }
                        new_sub_id = self.env['documents.folder'].create(folder_vals)
                        if new_sub_id:
                            self.write({field:new_sub_id})

        return True

class Pose(models.Model):
    _name = 'project.pose'
    _order = 'project_id, id desc'
    _inherit = 'documents.mixin'

    def _get_document_vals(self, attachment):
        self.ensure_one()
        vals = super(Pose,self)._get_document_vals(attachment)
        vals['res_model'] = 'project.project'
        vals['res_id'] = self.project_id.id
        return vals
    def _get_document_folder(self):
        return self.project_id.documents_folder_pose
    def _get_document_partner(self):
        return self.project_id.partner_id

    project_id = fields.Many2one('project.project', string='Projet', required=True)
    date_start_install = fields.Date("Date de début")
    date_end_install = fields.Date("Date de fin")
    notes = fields.Text(string='Notes')
    return_caution = fields.Boolean('Retour chq Caution', default=False)
    monitoring_mylight = fields.Boolean('Monitoring MyLight', default=False)
    enphase = fields.Boolean('Enphase', default=False)

    #PJs
    aft = fields.Many2many('ir.attachment', 'ir_attachment_pose_aft', string='AFT')
    picture = fields.Many2many('ir.attachment', 'ir_attachment_pose_picture', string='Photos')
    calepinage_emphase = fields.Many2many('ir.attachment', 'ir_attachment_pose_calepinage', string='Calepinage Enphase')
    implantation_emphase = fields.Many2many('ir.attachment', 'ir_attachment_pose_implantation', string='Rapport Enphase')
    quotation_alaska = fields.Many2many('ir.attachment', 'ir_attachment_pose_quot_alaska', string='Devis Alaska')
    invoice_alaska = fields.Many2many('ir.attachment', 'ir_attachment_pose_inv_alaska', string='Facture alaska')
    invoice_finalRsp = fields.Many2many('ir.attachment', 'ir_attachment_pose_inv_rsp', string='Facture finale RSP client')

class Fdi(models.Model):
    _name = 'fdi.object'
    _inherit = 'documents.mixin'
    def _get_document_vals(self, attachment):
        self.ensure_one()
        vals = super(Fdi,self)._get_document_vals(attachment)
        vals['res_model'] = 'project.project'
        vals['res_id'] = self.project_id.id
        return vals
    def _get_document_folder(self):
        return self.project_id.documents_folder_fdi
    def _get_document_partner(self):
        return self.project_id.partner_id

    state = fields.Selection([
        ('planif', 'Programmée'),
        ('finish', 'Terminée'),
    ], string="État")
    project_id = fields.Many2one('project.project', string='Projet', required=True, ondelete='cascade')
    aft_fdi = fields.Many2many('ir.attachment', 'ir_attachment_aft_fdi', string='AFT')
    date = fields.Datetime('Date')
    pictures_fdi = fields.Many2many('ir.attachment', 'ir_attachment_pictures_fdi', string='Photos')
    cause = fields.Char('Cause interruption')

    @api.onchange('date')
    def _on_change_state(self):
        if self.date:
            self.state = 'planif'

    def yes_finish(self):
        self.project_id.stage_id = self.env.ref('adquat_rsp.project_project_stage_mes').id
        self.state = 'finish'

class Sav(models.Model):
    _name = 'sav.object'
    _inherit = 'documents.mixin'
    def _get_document_vals(self, attachment):
        self.ensure_one()
        vals = super(Sav,self)._get_document_vals(attachment)
        vals['res_model'] = 'project.project'
        vals['res_id'] = self.project_id.id
        return vals
    def _get_document_folder(self):
        return self.project_id.documents_folder_sav
    def _get_document_partner(self):
        return self.project_id.partner_id

    project_id = fields.Many2one('project.project', required=True, ondelete='cascade')
    type_sav = fields.Selection([
        ('1', 'Toiture'),
        ('2', 'Elec'),
        ('3', 'Autre')
    ], string="Type de SAV")

    other_type_sav = fields.Char('Autre type de SAV')
    date = fields.Datetime('Date')
    return_picture = fields.Many2many('ir.attachment', 'ir_attachment_return_pic_sav', string='Retour Photo')
    sheet_intervention = fields.Many2many('ir.attachment', 'ir_attachment_sheet_inter_sav', string='Feuille d\'intervention')
    picture_sav = fields.Many2many('ir.attachment', 'ir_attachment_picture_sav', string='Photos SAV')
    state = fields.Selection([('planif', 'Programmé'),
        ('finish', 'Terminé'),], string="État")

    @api.onchange('date')
    def _on_change_state(self):
        if self.date:
            self.state = 'planif'

    def mise_en_service(self):
        self.state = 'finish'
        self.project_id.stage_id = self.env.ref('adquat_rsp.project_project_stage_mes').id

    def close_project(self):
        self.state = 'finish'
        self.project_id.stage_id = self.env.ref('adquat_rsp.project_project_stage_done').id
